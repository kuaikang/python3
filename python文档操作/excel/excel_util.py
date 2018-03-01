from openpyxl import load_workbook
from openpyxl.workbook import Workbook

def get_excel(filepath):
    '''返回excel中内容,以列表的形式返回'''
    try:
        wb = load_workbook(filepath)
        sheet_names = wb.sheetnames
        sheet = sheet_names[0]
        ws = wb[sheet]
        data_list = []
        for i in range(1,ws.max_row):
            list = []
            for j in range(1,ws.max_column):
                list.append(ws.cell(row=i,column=j).value)
            data_list.append(list)
        return data_list
    except Exception:
        print("读取excel文件发生错误")
        return None

def write_file(data,filename="a.txt"):
    with open(filename,mode="w",encoding="utf-8") as f:
        f.write(data)

def create_excel(list,file="a.xlsx"):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "sheet1"
    for i in range(len(list)):
        for j in range(len(list[i])):
            ws.cell(row=i+1,column=j+1,value=list[i][j])
    wb.save(file)

if __name__ == '__main__':
    data = get_excel("test.xlsx")
    print(data)
    create_excel(data)
