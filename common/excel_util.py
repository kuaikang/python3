from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import xlrd

def create_excel(data, file):
    """生成excel"""
    wb = Workbook()
    ws = wb.active
    length = len(data)
    for line in range(length):
        cols = len(data[line])
        for j in range(cols):
            ws.cell(row=line + 1, column=j + 1, value=data[line][j])
    wb.save(file)


def read_excel(filepath):
    '''返回excel中内容,以列表的形式返回'''
    try:
        data_list = []
        data = xlrd.open_workbook(filepath)
        table = data.sheets()[0]
        for i in range(0, table.nrows):
            values = table.row_values(i)  # 某一行数据
            data_list.append(values)
        return data_list
    except Exception:
        print("读取excel文件发生错误")
        return None
