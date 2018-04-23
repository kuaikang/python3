from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def get_excel(filepath):
    '''返回excel中内容,以列表的形式返回'''
    try:
        wb = load_workbook(filepath)
        sheet_names = wb.sheetnames
        sheet = sheet_names[0]
        ws = wb[sheet]
        data_list = []
        for i in range(1, ws.max_row):
            list = []
            for j in range(1, ws.max_column + 1):
                list.append(ws.cell(row=i, column=j).value)
            data_list.append(list)
        return data_list
    except Exception:
        print("读取excel文件发生错误")
        return None


def create_excel(data, file):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "sheet1"
    length = len(data)
    for i in range(length):
        for j in range(len(data[i])):
            ws.cell(row=i + 1, column=j + 1, value=data[i][j])
    wb.save(file)


if __name__ == '__main__':
    pass
